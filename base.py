# todo:

#make path both sample path and process path


import argparse
import pickle
import os
import re
from datetime import datetime
import sys
import ctypes
import subprocess
import tkinter as tk
from tkinter import filedialog
import shutil
#import openpyxl
from openpyxl import load_workbook
import win32com.client
from pathlib import Path

# Enable ANSI escape codes on Windows
kernel32 = ctypes.windll.kernel32
kernel32.SetConsoleMode(kernel32.GetStdHandle(-11), 7)

# Define ANSI color codes
RED = '\033[91m'
GREEN = '\033[92m'
BLUE = '\033[94m'
RESET = '\033[0m'
YELLOW = '\033[93m'
MAGENTA = '\033[95m'

IDbase_dir = r"I:\e24\SQN\Researchers\Haubmann Benjamin\01_PhD\IDbase"
SampleOverview_dir = r"I:\e24\SQN\Researchers\Haubmann Benjamin\01_PhD\Sample Overview.xlsx"
sheet_name = "Tabelle1"

IDdir_dic = {"sem": r"I:\e24\SQN\Researchers\Haubmann Benjamin\01_PhD\16_SEM",
             "plm": r"I:\e24\SQN\Researchers\Haubmann Benjamin\01_PhD\13_PL",
             "epi": r"I:\e24\SQN\Researchers\Haubmann Benjamin\01_PhD\15_Growth",
             "elx": r"I:\e24\SQN\Researchers\Haubmann Benjamin\01_PhD\12_Elionix",
             "mic": r"I:\e24\SQN\Researchers\Haubmann Benjamin\01_PhD\14_Microscope",
             "xrd": r"I:\e24\SQN\Researchers\Haubmann Benjamin\01_PhD\20_XRD",
             "tem": r"I:\e24\SQN\Researchers\Haubmann Benjamin\01_PhD\21_TEM",
             "mla": r"I:\e24\SQN\Researchers\Haubmann Benjamin\01_PhD\23_MLA",
             "rie": r"I:\e24\SQN\Researchers\Haubmann Benjamin\01_PhD\24_RIE"}

Sampledir_dic = {"sem": "SEM", "plm": "PL", "epi": "MBE", "elx": "Elionix", "mic": "Microscope", "xrd": "XRD", "tem": "TEM", "mla": "MLA", "rie": "RIE"}

SampleOverview_column_dic = {"sem": "S", "plm": "T", "epi": "P", "elx": "H", "mic": "R", "xrd": "U", "tem": "V", "mla": "I", "rie": "M"}


def extract_ID_from_path(path):
    ID = path.split("\\")[-1][:16]
    if os.path.isfile(path):
        path = "\\".join(path.split("\\")[:-1])
        print(path)
    return ID, path


def ID_exists(ID, base):
    if ID in base.keys():
        return True
    else:
        return False


def get_process_subdir(ID):
    for key in IDdir_dic.keys():
        if key in ID:
            return Sampledir_dic[key]
    return None


def get_sample_path(spl_name, base):
    for key in base.keys():
        if "spl" in key and spl_name in key:
            return base[key]["path"]
    return None


def sync_folder(ID):
    with open(IDbase_dir, 'rb') as file:
        base = pickle.load(file)

    if not ID_exists(ID, base):
        print(f"{RED}Invalid ID{RESET}")
        return

    tag_dict = base[ID].get("tags", {})
    if not tag_dict:
        return

    source = base[ID]["path"]
    count = 0
    for spl_name, copy_path in tag_dict.items():
        if os.path.exists(copy_path) and not os.path.isdir(copy_path):
            print(f"{YELLOW}Skipping \"{spl_name}\": copy path is a file, not a folder — please fix tag manually{RESET}")
            continue
        if os.path.isdir(copy_path):
            shutil.rmtree(copy_path)
        shutil.copytree(source, copy_path)
        count += 1

    print(f"{GREEN}Synced \"{ID}\" to {count} sample folder(s){RESET}")


def sync_all():
    with open(IDbase_dir, 'rb') as file:
        base = pickle.load(file)

    synced = 0
    for ID in base.keys():
        tag_dict = base[ID].get("tags", {})
        if tag_dict:
            source = base[ID]["path"]
            for spl_name, copy_path in tag_dict.items():
                if os.path.exists(copy_path) and not os.path.isdir(copy_path):
                    print(f"{YELLOW}Skipping \"{spl_name}\" for \"{ID}\": copy path is a file, not a folder — please fix tag manually{RESET}")
                    continue
                if os.path.isdir(copy_path):
                    shutil.rmtree(copy_path)
                shutil.copytree(source, copy_path)
                synced += 1

    basepath = r"I:\e24\SQN\Researchers\Haubmann Benjamin\01_PhD\11_Samples"
    cleaned = 0
    if os.path.exists(basepath):
        for sample_folder in os.listdir(basepath):
            sample_folder_path = os.path.join(basepath, sample_folder)
            if not os.path.isdir(sample_folder_path):
                continue
            for process_folder in os.listdir(sample_folder_path):
                process_folder_path = os.path.join(sample_folder_path, process_folder)
                if not os.path.isdir(process_folder_path):
                    continue
                for item in os.listdir(process_folder_path):
                    item_path = os.path.join(process_folder_path, item)
                    if not os.path.isfile(item_path):
                        continue
                    ID = item[:16]
                    if not ID_exists(ID, base):
                        continue
                    for spl_name, copy_path in base[ID].get("tags", {}).items():
                        if os.path.dirname(copy_path) == process_folder_path and os.path.isdir(copy_path):
                            os.remove(item_path)
                            cleaned += 1
                            break

    print(f"{GREEN}Synced {synced} copy/copies, removed {cleaned} loose file(s){RESET}")


def tag(ID, spl_name):
    with open(IDbase_dir, 'rb') as file:
        base = pickle.load(file)

    if not ID_exists(ID, base):
        print(f"{RED}Invalid ID{RESET}")
        return

    if "tags" not in base[ID]:
        base[ID]["tags"] = {}

    if spl_name in base[ID]["tags"]:
        print(f"{YELLOW}Tag \"{spl_name}\" already exists for \"{ID}\"{RESET}")
        return

    sample_path = get_sample_path(spl_name, base)
    if sample_path is None:
        print(f"{RED}Sample \"{spl_name}\" not found in database{RESET}")
        return

    process = get_process_subdir(ID)
    if process is None:
        print(f"{RED}Could not determine process type for \"{ID}\"{RESET}")
        return

    folder_name = os.path.basename(base[ID]["path"])
    copy_path = os.path.join(sample_path, process, folder_name)

    if os.path.exists(copy_path):
        print(f"{YELLOW}Copy already exists at \"{copy_path}\", registering without re-copying{RESET}")
    else:
        shutil.copytree(base[ID]["path"], copy_path)

    base[ID]["tags"][spl_name] = copy_path

    with open(IDbase_dir, 'wb') as file:
        pickle.dump(base, file)

    print(f"{GREEN}Tagged \"{ID}\" with \"{spl_name}\" → {copy_path}{RESET}")


def untag(ID, spl_name):
    with open(IDbase_dir, 'rb') as file:
        base = pickle.load(file)

    if not ID_exists(ID, base):
        print(f"{RED}Invalid ID{RESET}")
        return

    tag_dict = base[ID].get("tags", {})
    if spl_name not in tag_dict:
        print(f"{RED}Tag \"{spl_name}\" not found for \"{ID}\"{RESET}")
        return

    copy_path = tag_dict[spl_name]
    print(f"{BLUE}Delete the copy at \"{copy_path}\"? {GREEN}y{BLUE}/{RED}n{RESET}")
    choice = input()
    if choice == "y":
        if os.path.isdir(copy_path):
            shutil.rmtree(copy_path)
            print(f"{GREEN}Deleted copy{RESET}")
        elif os.path.isfile(copy_path):
            os.remove(copy_path)
            print(f"{GREEN}Deleted file{RESET}")
        else:
            print(f"{YELLOW}Copy path does not exist, skipping deletion{RESET}")
    elif choice == "n":
        pass
    else:
        print(f"{RED}Invalid input, tag not removed{RESET}")
        return

    del base[ID]["tags"][spl_name]

    with open(IDbase_dir, 'wb') as file:
        pickle.dump(base, file)

    print(f"{GREEN}Tag \"{spl_name}\" removed from \"{ID}\"{RESET}")



def untagged():
    with open(IDbase_dir, 'rb') as file:
        base = pickle.load(file)

    exclude = {"sim", "scr", "ana", "spl", "des"}
    result = []
    for ID in base.keys():
        if any(ex in ID for ex in exclude):
            continue
        if not base[ID].get("tags"):
            result.append(ID)

    if not result:
        print(f"{GREEN}All IDs have at least one tag{RESET}")
    else:
        for ID in sorted(result):
            print(f"{MAGENTA}{ID}{RESET}")



def list_tags(ID):
    with open(IDbase_dir, 'rb') as file:
        base = pickle.load(file)

    if not ID_exists(ID, base):
        print(f"{RED}Invalid ID{RESET}")
        return

    tag_dict = base[ID].get("tags", {})
    if not tag_dict:
        print(f"{YELLOW}No tags for \"{ID}\"{RESET}")
    else:
        for spl_name, copy_path in tag_dict.items():
            status = f"{GREEN}ok{RESET}" if os.path.exists(copy_path) else f"{RED}missing{RESET}"
            print(f"  {MAGENTA}{spl_name}{RESET} → {copy_path} [{status}]")


class entry:

    def __init__(self, path):

        if not os.path.exists(path):
            print(f"{RED}Path does not exist{RESET}")
            return

        #self.path = path
        self.ID, self.path = extract_ID_from_path(path)

        with open(IDbase_dir, 'rb') as file:
            base = pickle.load(file)

        print(f"{BLUE}Please provide a short description about the entry you want to add{RESET}")
        info = input()
        separator = "#"*70

        with open(f"{self.path}\\{self.ID}_readme.txt", "a") as readme:
            readme.write(f"{info}\n\n{separator}\n\n")

        base[self.ID] = {"path": self.path, "info": info, "comments": "", "tags": {}}

        with open(IDbase_dir, 'wb') as file:
            pickle.dump(base, file)

        print(f"{GREEN}Entry \"{self.ID}\" has been added{RESET}")

        if "des" in self.ID:
            print(f"{BLUE}Paste the creation code into the editor window and save when done{RESET}")
            subprocess.run(['notepad', f"{self.path}\\{self.ID}_readme.txt"])
            update_readme_single(self.ID)

        excluded = {"des", "sim", "scr", "ana", "spl"}
        if not any(ex in self.ID for ex in excluded):
            while True:
                print(f"{BLUE}Which samples are involved? Enter sample names separated by commas (e.g. spl01,spl02), or press Enter to skip:{RESET}")
                sample_input = input().strip()
                if sample_input:
                    for spl_name in [s.strip() for s in sample_input.split(",") if s.strip()]:
                        tag(self.ID, spl_name)
                    break
                else:
                    print(f"{BLUE}No samples entered. Confirm no sample was involved? {GREEN}y{BLUE}/{RED}n{RESET}")
                    confirm = input().strip()
                    if confirm == "y":
                        break


def new_sample(spl_name):

    basepath = r"I:\e24\SQN\Researchers\Haubmann Benjamin\01_PhD\11_Samples"
    current_date = datetime.now()
    date = current_date.strftime('%Y%m%d')

    with open(IDbase_dir, 'rb') as file:
        base = pickle.load(file)

    samples = sorted([key for key in base.keys() if "spl" in key])

    if any([True for x in samples if spl_name in x]):
        print(f"{RED}Sample already exists.{RESET}")
        return

    else:
        path = basepath + "\\" + date + "-" + spl_name
        os.makedirs(path)
        for key in Sampledir_dic:
            subfolder_name = Sampledir_dic[key]
            os.makedirs(path + "\\" + subfolder_name)

        add(path)


def convert_date_format(date_str):

    if len(date_str) != 8 or not date_str.isdigit():
        raise ValueError("Input must be a valid date in 'YYYYMMDD' format.")

    year = date_str[:4]
    month = date_str[4:6]
    day = date_str[6:8]

    return f"{day}.{month}.{year}"


def write_to_cell(file_name, sheet_name, cell_address, value):

    workbook = load_workbook(file_name)
    sheet = workbook[sheet_name]

    existing_value = sheet[cell_address].value
    if existing_value == None:
        sheet[cell_address] = value
    else:
        sheet[cell_address] = (str(sheet[cell_address].value) or "") + "\n\n" + value
    workbook.save(file_name)
    print(f"Value '{value}' written to {cell_address} in sheet '{sheet_name}' of '{file_name}'.")

def save_close_excel(file_path):

    xl = win32com.client.GetActiveObject("Excel.Application")

    # Look for the workbook in open workbooks
    for wb in xl.Workbooks:
        if wb.FullName == file_path:
            print(f"Sample Overview active")
            wb.Save()  # Save before closing
            wb.Close()
            print(f"Closed Sample Overview")
            break



def reopen_excel(file_path):

    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = True  # Open Excel in visible mode
        excel.Workbooks.Open(file_path)
        print(f"Opened Sample Overview")
    except Exception as e:
        print(f"Error while opening Sample Overview")


def get_sample_index(spl_name):

    with open(IDbase_dir, 'rb') as file:
        base = pickle.load(file)

    samples = sorted([key for key in base.keys() if "spl" in key])

    index = 0
    for i, sample in enumerate(samples):
        if spl_name in sample:
            return index
        else:
            index += 1
    print("No matching sample found")


def update_SampleOverview(ID, spl):

    date = convert_date_format(ID[:8])
    print(get_sample_index(spl))
    row = str(get_sample_index(spl) + 2)
    print(row)
    column = get_column(ID)
    cell = column + row
    value = f"{date}, {ID}"

    save_close_excel(SampleOverview_dir)

    write_to_cell(SampleOverview_dir, sheet_name, cell, value)

    # In case of Elionix process: Write design-ID in column "R"
    if "elx" in ID:
        design_ID = input("Design-ID: ")
        design_cell = "R" + row
        write_to_cell(SampleOverview_dir, sheet_name, design_cell, design_ID)

    reopen_excel(SampleOverview_dir)


def get_column(ID):
    for key in IDdir_dic.keys():
        if key in ID:
            column = SampleOverview_column_dic[key]
    return column


def create(new_name, initial_path = r"I:\e24\SQN\Researchers\Haubmann Benjamin\01_PhD"):

    root = tk.Tk()
    root.withdraw()

    folder_path = filedialog.askdirectory(title="Select a Folder", initialdir=initial_path)

    parent_dir = os.path.dirname(folder_path)

    renamed_folder_path = os.path.join(parent_dir, new_name)
    os.rename(folder_path, renamed_folder_path)

    ID = new_name[:16]

    if len(new_name) > 16:
        description = new_name[16:]

    for key in IDdir_dic.keys():
        if key in ID:
            new_parent_dir = IDdir_dic[key]

    new_path = shutil.move(renamed_folder_path, new_parent_dir)

    add(new_path)



def add(path):
    entry(path)



def goto(ID):

    with open(IDbase_dir, 'rb') as file:
        base = pickle.load(file)

    if not ID_exists(ID, base):
        print(f"{RED}Invalid ID{RESET}")
        return

    path = base[ID]["path"]
    os.startfile(path)


def delete(ID):

    with open(IDbase_dir, 'rb') as file:
        base = pickle.load(file)

    if not ID_exists(ID, base):
        print(f"{RED}Invalid ID{RESET}")
        return

    tag_dict = base[ID].get("tags", {})

    print(f"{BLUE}Do you really want to delete entry \"{ID}\"? {GREEN}y{BLUE}/{RED}n{RESET}")
    choice = input()
    if choice == "y":

        deleted_copy_paths = set()
        if tag_dict:
            print(f"{BLUE}Also delete {len(tag_dict)} sample folder copy/copies? {GREEN}y{BLUE}/{RED}n{RESET}")
            copy_choice = input()
            if copy_choice == "y":
                for spl_name, copy_path in tag_dict.items():
                    if os.path.exists(copy_path):
                        shutil.rmtree(copy_path)
                        deleted_copy_paths.add(copy_path)
                        print(f"{GREEN}Deleted copy for \"{spl_name}\"{RESET}")

        print(f"{BLUE}Also delete the readme file? {GREEN}y{BLUE}/{RED}n{RESET}")
        readme_choice = input()
        if readme_choice == "y":
            readme_path = base[ID]["path"] + "\\" + ID + "_readme.txt"
            if os.path.exists(readme_path):
                os.remove(readme_path)
                print(f"{GREEN}Deleted readme from main path{RESET}")
            for spl_name, copy_path in tag_dict.items():
                if copy_path in deleted_copy_paths:
                    continue
                tagged_readme = copy_path + "\\" + ID + "_readme.txt"
                if os.path.exists(tagged_readme):
                    os.remove(tagged_readme)
                    print(f"{GREEN}Deleted readme from \"{spl_name}\" copy{RESET}")

        del base[ID]
        with open(IDbase_dir, 'wb') as file:
            pickle.dump(base, file)
        print(f"{GREEN}Entry \"{ID}\" has been deleted{RESET}")

    elif choice == "n":
        return

    else:
        print(f"{RED}Invalid Entry{RESET}")
        delete(ID)



def ls():
    with open(IDbase_dir, 'rb') as file:
        base = pickle.load(file)
    l = []
    for ID in base.keys():
        l.append(ID)

    for ID in sorted(l):
        print(f"{MAGENTA}{ID}{RESET}")



def checkall():

    with open(IDbase_dir, 'rb') as file:
        base = pickle.load(file)

    invalid_list = [ID for ID in base.keys() if not os.path.exists(base[ID]["path"])]

    if len(invalid_list) == 0:
        print(f"{GREEN}Everything seems up to date{RESET}")
        return

    print(f"{RED}Invalid path found for {len(invalid_list)} entrie(s). Searching...{RESET}\n")

    for ID in invalid_list:
        expected_dir = None
        for key in IDdir_dic.keys():
            if key in ID:
                expected_dir = IDdir_dic[key]
                break

        if expected_dir is None or not os.path.exists(expected_dir):
            print(f"{RED}{ID}{RESET} — could not determine expected directory, skipping")
            continue

        matches = [f for f in os.listdir(expected_dir) if f[:16] == ID and os.path.isdir(os.path.join(expected_dir, f))]

        if not matches:
            print(f"{RED}{ID}{RESET} — not found in {expected_dir}, skipping")
            continue

        for match in matches:
            new_path = os.path.join(expected_dir, match)
            print(f"{YELLOW}{ID}{RESET} — found at: {new_path}")
            print(f"{BLUE}Update path? {GREEN}y{BLUE}/{RED}n{RESET}")
            choice = input()
            if choice == "y":
                base[ID]["path"] = new_path
                with open(IDbase_dir, 'wb') as file:
                    pickle.dump(base, file)
                with open(IDbase_dir, 'rb') as file:
                    base = pickle.load(file)
                print(f"{GREEN}Path updated{RESET}")
            else:
                print(f"{YELLOW}Skipped{RESET}")


def update(path):

    if not os.path.exists(path):
        print(f"{RED}Path does not exist{RESET}")
        return

    ID, path = extract_ID_from_path(path)

    with open(IDbase_dir, 'rb') as file:
        base = pickle.load(file)

    if not ID in base.keys():
        print(f"{RED}Path exists, but ID has not been added yet{RESET}")
        return

    oldpath = base[ID]["path"]
    print(f"{BLUE}Do you really want to change the path of entry \"{ID}\" from \"{oldpath}\" to \"{path}\"? {GREEN}y{BLUE}{RED}n{RESET}")

    choice = input()
    if choice == "y":

        base[ID]["path"] = path
        with open(IDbase_dir, 'wb') as file:
            pickle.dump(base, file)
        print(f"{GREEN}Path has been updated{RESET}")
        sync_folder(ID)

    elif choice == "n":
        return

    else:
        print(f"{RED}Invalid Entry{RESET}")
        update(path)


def update_readme_single(ID):

    with open(IDbase_dir, 'rb') as file:
        base = pickle.load(file)

    readme_path = base[ID]["path"] + "\\" + ID + "_readme.txt"

    if not os.path.exists(readme_path):
        print(f"{YELLOW}Skipping \"{ID}\": readme not found{RESET}")
        return False

    with open(readme_path, 'r') as readme:
        content = readme.read()

    parts = [y.strip() for y in content.split("#"*70, 1)]
    if len(parts) != 2:
        print(f"{YELLOW}Skipping \"{ID}\": readme has no separator, cannot parse{RESET}")
        return False

    info_new, comments_new = parts
    base[ID]["info"] = info_new
    base[ID]["comments"] = comments_new

    with open(IDbase_dir, 'wb') as file:
        pickle.dump(base, file)

    sync_folder(ID)
    return True



def update_readme():

    with open(IDbase_dir, 'rb') as file:
        base = pickle.load(file)

    ok = 0
    skipped = 0
    for ID in base.keys():
        if update_readme_single(ID):
            ok += 1
        else:
            skipped += 1

    print(f"{GREEN}Updated {ok} readme(s){RESET}" + (f", skipped {skipped}" if skipped else ""))



def display(type_):

    with open(IDbase_dir, 'rb') as file:
        base = pickle.load(file)

    l = []
    for ID in base.keys():
        l.append(ID)

    for ID in sorted(l):
        if type_ in ID or type_ == "all":
            info = base[ID]["info"]
            tag_dict = base[ID].get("tags", {})
            tag_str = f" {YELLOW}[{', '.join(tag_dict.keys())}]{RESET}" if tag_dict else ""
            print(f"{MAGENTA}{ID}{RESET}{tag_str}\n{info}\n")


def comment(ID):

    with open(IDbase_dir, 'rb') as file:
        base = pickle.load(file)

    if not ID in base.keys():
        print(f"{RED}Invalid ID{RESET}")
        return

    print(f"{BLUE}Type your comment{RESET}")

    comment = input()
    now = datetime.now()
    full_date = now.strftime("%A, %B %d, %Y")
    comment = f"{full_date}:\n{comment}\n\n"

    with open(base[ID]["path"]+"\\"+ID+"_readme.txt", 'a') as readme:
        readme.write(comment)

    update_readme_single(ID)

    print(f"{GREEN}Comment added{RESET}")


def inspect(ID):

    with open(IDbase_dir, 'rb') as file:
        base = pickle.load(file)

    if not ID in base.keys():
        print(f"{RED}Invalid ID{RESET}")
        return

    with open(base[ID]["path"]+"\\"+ID+"_readme.txt", 'r') as readme:
        content = readme.read()

    print(content)


def edit_readme(ID):

    with open(IDbase_dir, 'rb') as file:
        base = pickle.load(file)

    if not ID_exists(ID, base):
        print(f"{RED}Invalid ID{RESET}")
        return

    readme_path = base[ID]["path"] + "\\" + ID + "_readme.txt"

    if not os.path.exists(readme_path):
        print(f"{RED}Readme not found: {readme_path}{RESET}")
        return

    subprocess.run(['notepad', readme_path])

    update_readme_single(ID)
    print(f"{GREEN}Readme saved and synced{RESET}")





def _get_excel_row(row_num):
    """Return (headers, row_data) dicts keyed by column letter, or (None, None) on failure."""
    paths_to_try = [
        SampleOverview_dir,
        r"\\nas.ads.mwn.de\tuze\wsi\e24\SQN\Researchers\Haubmann Benjamin\01_PhD\Sample Overview Local.xlsx",
        r"C:\Users\ge23jud\OneDrive - TUM\Sample Overview.xlsx",
    ]
    for path in paths_to_try:
        try:
            wb = load_workbook(path, data_only=True)
            ws = wb[sheet_name]
            headers = {cell.column_letter: cell.value for cell in ws[1] if cell.value is not None}
            row_data = {cell.column_letter: cell.value for cell in ws[row_num] if cell.value is not None}
            return headers, row_data
        except Exception:
            continue
    return None, None


def info(query):
    with open(IDbase_dir, 'rb') as file:
        base = pickle.load(file)

    matches = [ID for ID in base.keys() if query in ID]

    if not matches:
        print(f"{RED}No entry found matching \"{query}\"{RESET}")
        return

    samples_sorted = sorted(k for k in base.keys() if "spl" in k)

    if "spl" in query:
        for spl_ID in sorted(matches):
            entry_data = base[spl_ID]
            print(f"\n{BLUE}{'='*60}{RESET}")
            print(f"{MAGENTA}{spl_ID}{RESET}")
            print(f"{BLUE}{'='*60}{RESET}")
            print(f"{YELLOW}Path:{RESET} {entry_data['path']}")
            print(f"\n{YELLOW}Info:{RESET}\n{entry_data['info']}")
            if entry_data.get('comments', '').strip():
                print(f"\n{YELLOW}Comments:{RESET}\n{entry_data['comments']}")

            tagged_processes = []
            for ID, data in base.items():
                if "spl" in ID:
                    continue
                for tag_key in data.get("tags", {}):
                    if query in tag_key:
                        tagged_processes.append((ID, data, data["tags"][tag_key]))
                        break

            if tagged_processes:
                print(f"\n{YELLOW}Tagged processes ({len(tagged_processes)}):{RESET}")
                for proc_ID, proc_data, copy_path in sorted(tagged_processes):
                    process_type = get_process_subdir(proc_ID) or "?"
                    path_status = f"{GREEN}ok{RESET}" if os.path.exists(proc_data['path']) else f"{RED}missing{RESET}"
                    print(f"\n  {MAGENTA}{proc_ID}{RESET} ({process_type}) [{path_status}]")
                    print(f"  {YELLOW}Path:{RESET} {proc_data['path']}")
                    if proc_data['info'].strip():
                        print(f"  {YELLOW}Info:{RESET} {proc_data['info']}")
            else:
                print(f"\n{YELLOW}No processes tagged to this sample{RESET}")

            spl_idx = next((i for i, k in enumerate(samples_sorted) if query in k), None)
            if spl_idx is not None:
                headers, row_data = _get_excel_row(spl_idx + 2)
                if headers is not None and row_data:
                    col_by_header = {v: k for k, v in headers.items()}

                    def _epi_from_col(rd, col_name):
                        col = col_by_header.get(col_name)
                        if col and col in rd:
                            m = re.search(r'(?i)epi[-]?(\d+)', str(rd[col]))
                            return f"EPI-{m.group(1)}" if m else None
                        return None

                    growth_origin = []

                    epi = _epi_from_col(row_data, "Growth")
                    if epi:
                        growth_origin.append(f"{epi} (sample is growth wafer)")

                    nw_col = col_by_header.get("NW Transfer")
                    if nw_col and nw_col in row_data:
                        nw_val = str(row_data[nw_col])
                        if 'from' in nw_val.lower():
                            m = re.search(r'(?i)epi[-]?(\d+)', nw_val)
                            if m:
                                growth_origin.append(f"EPI-{m.group(1)} (NWs transferred from)")

                    cleaved_col = col_by_header.get("Cleaved From")
                    if cleaved_col and cleaved_col in row_data:
                        parent = str(row_data[cleaved_col]).strip()
                        if parent.lower() not in ('unknown', ''):
                            parent_idx = next((i for i, k in enumerate(samples_sorted) if parent in k), None)
                            if parent_idx is not None:
                                p_headers, p_row = _get_excel_row(parent_idx + 2)
                                if p_headers and p_row:
                                    p_col_by_header = {v: k for k, v in p_headers.items()}
                                    p_growth_col = p_col_by_header.get("Growth")
                                    if p_growth_col and p_growth_col in p_row:
                                        m = re.search(r'(?i)epi[-]?(\d+)', str(p_row[p_growth_col]))
                                        if m:
                                            growth_origin.append(f"EPI-{m.group(1)} (cleaved from {parent})")

                    keep = {"Type", "eSAE", "Ellipsometry", "Clean", "Spin-Coating",
                            "Development", "HF 1", "HF 2", "NW Transfer", "Design", "Cleaved"}
                    print(f"\n{YELLOW}Sample Overview:{RESET}")
                    for origin in growth_origin:
                        print(f"  {YELLOW}Growth origin:{RESET} {origin}")
                    for col_letter, value in row_data.items():
                        col_name = headers.get(col_letter, col_letter)
                        if col_name in keep:
                            display = ' | '.join(p.strip() for p in str(value).split('\n\n') if p.strip())
                            print(f"  {YELLOW}{col_name}:{RESET} {display}")
                elif headers is None:
                    print(f"\n{YELLOW}Sample Overview not accessible{RESET}")
    else:
        for proc_ID in sorted(matches):
            entry_data = base[proc_ID]
            process_type = get_process_subdir(proc_ID) or "?"
            path_status = f"{GREEN}ok{RESET}" if os.path.exists(entry_data['path']) else f"{RED}missing{RESET}"
            print(f"\n{BLUE}{'='*60}{RESET}")
            print(f"{MAGENTA}{proc_ID}{RESET} ({process_type}) [{path_status}]")
            print(f"{BLUE}{'='*60}{RESET}")
            print(f"{YELLOW}Path:{RESET} {entry_data['path']}")
            print(f"\n{YELLOW}Info:{RESET}\n{entry_data['info']}")
            if entry_data.get('comments', '').strip():
                print(f"\n{YELLOW}Comments:{RESET}\n{entry_data['comments']}")

            tag_dict = entry_data.get("tags", {})
            if tag_dict:
                print(f"\n{YELLOW}Tagged to:{RESET}")
                for spl_name, copy_path in tag_dict.items():
                    copy_status = f"{GREEN}ok{RESET}" if os.path.exists(copy_path) else f"{RED}missing{RESET}"
                    print(f"  {MAGENTA}{spl_name}{RESET} → {copy_path} [{copy_status}]")
            else:
                print(f"\n{YELLOW}Not tagged to any sample{RESET}")

            proc_key = next((k for k in IDdir_dic if k in proc_ID), None)
            excel_col = SampleOverview_column_dic.get(proc_key)
            if excel_col and tag_dict:
                excel_entries = []
                for spl_name in tag_dict:
                    spl_idx = next((i for i, k in enumerate(samples_sorted) if spl_name in k), None)
                    if spl_idx is not None:
                        headers, row_data = _get_excel_row(spl_idx + 2)
                        if headers is not None and excel_col in row_data:
                            col_name = headers.get(excel_col, excel_col)
                            excel_entries.append((spl_name, col_name, row_data[excel_col]))
                if excel_entries:
                    print(f"\n{YELLOW}Sample Overview:{RESET}")
                    for spl_name, col_name, cell_val in excel_entries:
                        print(f"  {MAGENTA}{spl_name}{RESET} — {col_name}: {str(cell_val).strip()}")

            if "epi" in proc_ID:
                epi_m = re.search(r'epi[-]?(\d+)', proc_ID)
                if epi_m:
                    epi_num = epi_m.group(1)
                    grown_spls = set()
                    cleaved_map = {}
                    nw_epi_spls = []
                    nw_spl_map = {}
                    for i, spl_key in enumerate(samples_sorted):
                        spl_m = re.search(r'spl(\d+)', spl_key)
                        if not spl_m:
                            continue
                        spl_short = "spl" + spl_m.group(1)
                        hdrs, rdata = _get_excel_row(i + 2)
                        if hdrs is None:
                            continue
                        cbh = {v: k for k, v in hdrs.items()}
                        g_col = cbh.get("Growth")
                        if g_col and g_col in rdata and re.search(rf'(?i)epi[-]?{epi_num}\b', str(rdata[g_col])):
                            grown_spls.add(spl_short)
                        c_col = cbh.get("Cleaved From")
                        if c_col and c_col in rdata:
                            parent_val = str(rdata[c_col]).strip()
                            if parent_val and parent_val.lower() != 'unknown':
                                cleaved_map[spl_short] = parent_val
                        nw_col = cbh.get("NW Transfer")
                        if nw_col and nw_col in rdata:
                            nw_val = str(rdata[nw_col])
                            if 'from' in nw_val.lower():
                                if re.search(rf'(?i)epi[-]?{epi_num}\b', nw_val):
                                    nw_epi_spls.append(spl_short)
                                else:
                                    nw_src_m = re.search(r'spl\d+', nw_val, re.IGNORECASE)
                                    if nw_src_m:
                                        nw_spl_map[spl_short] = nw_src_m.group(0).lower()
                    seen_d = set(tag_dict.keys())
                    derived = []
                    for s in nw_epi_spls:
                        if s not in seen_d:
                            derived.append((s, "NW transfer from this wafer"))
                            seen_d.add(s)
                    cleaved_pieces = set()
                    for s, parent_val in cleaved_map.items():
                        parent_short_m = re.search(r'spl\d+', parent_val, re.IGNORECASE)
                        if parent_short_m and parent_short_m.group(0).lower() in {g.lower() for g in grown_spls}:
                            cleaved_pieces.add(s)
                            if s not in seen_d:
                                derived.append((s, f"cleaved from {parent_val}"))
                                seen_d.add(s)
                    for s, src in nw_spl_map.items():
                        if s not in seen_d and src in {c.lower() for c in cleaved_pieces}:
                            src_parent = cleaved_map.get(src, "?")
                            derived.append((s, f"NW transfer from {src} (cleaved from {src_parent})"))
                            seen_d.add(s)
                    if derived:
                        print(f"\n{YELLOW}Derived samples:{RESET}")
                        for s, rel in sorted(derived):
                            print(f"  {MAGENTA}{s}{RESET} — {rel}")


def parse_arguments():
    parser = argparse.ArgumentParser(description='Call individual functions from the command line.')
    subparsers = parser.add_subparsers(dest='function', required=True)

    # Subparser for add
    parser_add = subparsers.add_parser('add', help='Call add')
    parser_add.add_argument('path', type=str, help='folder path of the entry to add')

    # Subparser for goto
    parser_goto = subparsers.add_parser("goto", help="Call goto")
    parser_goto.add_argument("ID", type=str, help="ID which should be opened")

    # Subparser for ls
    parser_ls = subparsers.add_parser('ls', help='Call ls')

    # Subparser for delete
    parser_delete = subparsers.add_parser("delete", help="Call delete")
    parser_delete.add_argument("ID", type=str, help="ID which should be deleted")

    # Subparser for checkall
    parser_checkall = subparsers.add_parser('checkall', help='Call checkall')

    # Subparser for update
    parser_update = subparsers.add_parser("update", help="Call update")
    parser_update.add_argument("path", type=str, help="path which should be updated")

    # Subparser for display
    parser_display = subparsers.add_parser("display", help="Call display")
    parser_display.add_argument("type_", type=str, help="type of entries to be displayed")

    # Subparser for update_readme
    parser_update_readme = subparsers.add_parser('update_readme', help='Call update_readme')

    # Subparser for comment
    parser_comment = subparsers.add_parser("comment", help="Call comment")
    parser_comment.add_argument("ID", type=str, help="Comment to add")

    # Subparser for inspect
    parser_inspect = subparsers.add_parser("inspect", help="Call inspect")
    parser_inspect.add_argument("ID", type=str, help="ID whose readme is to show")

    # Subparser for edit_readme
    parser_edit_readme = subparsers.add_parser("edit_readme", help="Open readme in text editor and sync on close")
    parser_edit_readme.add_argument("ID", type=str, help="ID whose readme to edit")

    # Subparser for create
    parser_create = subparsers.add_parser("create", help="Create ID dir from folder and sort automatically")
    parser_create.add_argument("new_name", type=str, help="name of the ID folder")

    # Subparser for new_sample
    parser_new_sample = subparsers.add_parser("new_sample", help="Create new sample")
    parser_new_sample.add_argument("spl_name", type=str, help="Sample name")

    # Subparser for reopen_excel
    parser_reopen_excel = subparsers.add_parser("reopen_excel", help="Open Sample Overview")
    parser_reopen_excel.add_argument("file_path", type=str, help="Path to Sample Overview")

    # Subparser for write_to_cell
    parser_write_to_cell = subparsers.add_parser("write_to_cell", help="Write value to excel cell")
    parser_write_to_cell.add_argument("file_name", type=str, help="Name of excel file")
    parser_write_to_cell.add_argument("sheet_name", type=str, help="Name of excel sheet")
    parser_write_to_cell.add_argument("cell_address", type=str, help="Which cell to write to")
    parser_write_to_cell.add_argument("value", type=str, help="What to write to cell")

    # Subparser for save_close_excel
    parser_save_close_excel = subparsers.add_parser("save_close_excel", help="Save and close Sample Overview")
    parser_save_close_excel.add_argument("file_path", type=str, help="Path to Sample Overview")

    # Subparser for update_SampleOverview
    parser_update_SampleOverview = subparsers.add_parser("update_SampleOverview", help="Update SampleOverview")
    parser_update_SampleOverview.add_argument("ID", type=str, help="process ID")
    parser_update_SampleOverview.add_argument("spl", type=str, help="sample number")

    # Subparser for tag
    parser_tag = subparsers.add_parser("tag", help="Tag an ID with a sample")
    parser_tag.add_argument("ID", type=str, help="Process ID to tag")
    parser_tag.add_argument("spl_name", type=str, help="Sample name to tag with")

    # Subparser for untag
    parser_untag = subparsers.add_parser("untag", help="Remove a sample tag from an ID")
    parser_untag.add_argument("ID", type=str, help="Process ID to untag")
    parser_untag.add_argument("spl_name", type=str, help="Sample name to remove")

    # Subparser for sync
    parser_sync = subparsers.add_parser("sync", help="Sync ID folder to all tagged sample copies")
    parser_sync.add_argument("ID", type=str, help="ID to sync")

    # Subparser for sync_all
    parser_sync_all = subparsers.add_parser("sync_all", help="Sync all tagged ID folders to their sample copies")

    # Subparser for tags
    parser_tags = subparsers.add_parser("tags", help="List tags for an ID")
    parser_tags.add_argument("ID", type=str, help="ID whose tags to list")

    # Subparser for untagged
    parser_untagged = subparsers.add_parser("untagged", help="List all IDs without tags (excludes sim, scr, ana)")

    # Subparser for info
    parser_info = subparsers.add_parser("info", help="Show all information about a sample or process entry")
    parser_info.add_argument("query", type=str, help="Sample name (e.g. spl2407) or process number (e.g. epi1780)")

    return parser.parse_args()

if __name__ == "__main__":

    args = parse_arguments()

    if args.function == 'add':
        add(args.path)
    elif args.function == 'goto':
        goto(args.ID)
    elif args.function == "delete":
        delete(args.ID)
    elif args.function == "ls":
        ls()
    elif args.function == "checkall":
        checkall()
    elif args.function == "update":
        update(args.path)
    elif args.function == "display":
        display(args.type_)
    elif args.function == "update_readme":
        update_readme()
    elif args.function == "comment":
        comment(args.ID)
    elif args.function == "inspect":
        inspect(args.ID)
    elif args.function == "edit_readme":
        edit_readme(args.ID)
    elif args.function == "create":
        create(args.new_name)
    elif args.function == "new_sample":
        new_sample(args.spl_name)
    elif args.function == "reopen_excel":
        reopen_excel(args.file_path)
    elif args.function == "write_to_cell":
        write_to_cell(args.file_name, args.sheet_name, args.cell_address, args.value)
    elif args.function == "save_close_excel":
        save_close_excel(args.file_path)
    elif args.function == "update_SampleOverview":
        update_SampleOverview(args.ID, args.spl)
    elif args.function == "tag":
        tag(args.ID, args.spl_name)
    elif args.function == "untag":
        untag(args.ID, args.spl_name)
    elif args.function == "sync":
        sync_folder(args.ID)
    elif args.function == "sync_all":
        sync_all()
    elif args.function == "tags":
        list_tags(args.ID)
    elif args.function == "untagged":
        untagged()
    elif args.function == "info":
        info(args.query)



    #things to add

# edit path; check if path exists; search for Entry automatically
# When ID path is updated, readme file should move as well
